#!/usr/bin/env python3
"""
pdf_converter.py
================
Converts a Wazuh report file (XLSX or CSV) to a formatted PDF using fpdf2.

Works on Windows and Linux with no system-level dependencies -- fpdf2 is a
pure-Python library that embeds its own font handling.

The generated PDF:
  - Landscape A4 page
  - Wazuh-branded header (dark blue, configurable title and subtitle)
  - Auto-sized columns distributed across the page width
  - Full cell content always visible -- each row height is measured per-row
    so rows are only as tall as needed and follow continuously with no gaps
  - Alternating row shading for readability
  - Footer with page number and generation timestamp
  - Column headers repeated on every new page

Row rendering strategy
----------------------
fpdf2's multi_cell() adds inter-cell spacing that produces gaps between rows.
Instead, each row is rendered in two passes using cell():

  Pass 1 -- measure: iterate cells, count wrapped lines per cell using
            pdf.get_string_width(), take the maximum line count across
            all cells to get the row height.

  Pass 2 -- draw: render each cell at the measured row height using
            cell() with a manual word-wrap simulation (split the value
            into lines, draw each line as a sub-cell stacked vertically,
            fill the remainder of the cell height with a blank cell).

This produces tight, continuous rows with no inter-row spacing.

Dependencies:
  pip install fpdf2 openpyxl

Usage (standalone):
  python3 pdf_converter.py input.xlsx output.pdf --title "My Report"
  python3 pdf_converter.py input.csv  output.pdf --title "My Report"
"""

from __future__ import annotations

import argparse
import csv
import logging
from datetime import datetime
from pathlib import Path

# -- Colour palette -------------------------------------------------------------

HEADER_BG   = (26,  74, 107)
HEADER_FG   = (255, 255, 255)
ROW_EVEN    = (245, 247, 250)
ROW_ODD     = (255, 255, 255)
BORDER      = (200, 210, 220)
TITLE_FG    = (26,  74, 107)
SUBTITLE_FG = (100, 116, 135)
FOOTER_FG   = (150, 160, 170)


# -- Latin-1 sanitiser ---------------------------------------------------------

def _safe(text: str) -> str:
    """Replace characters outside Latin-1 with '?' -- Helvetica is Latin-1 only."""
    return "".join(c if ord(c) <= 255 else "?" for c in text)


# -- Data loading --------------------------------------------------------------

def load_xlsx(path: Path) -> tuple[list[str], list[list]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data    = [[str(c) if c is not None else "" for c in row] for row in rows[1:]]
    return headers, data


def load_csv(path: Path) -> tuple[list[str], list[list]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(f, dialect))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def load_file(path: Path) -> tuple[list[str], list[list]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_xlsx(path)
    elif suffix == ".csv":
        return load_csv(path)
    raise ValueError(f"Unsupported file type: '{suffix}'. Expected .xlsx or .csv")


# -- Column width calculation ---------------------------------------------------

def calculate_col_widths(
    headers: list[str],
    rows: list[list],
    page_width: float,
    font_size: float,
    max_sample: int = 100,
) -> list[float]:
    """
    Proportional column widths based on actual content length.
    No character cap -- columns reflect real data widths.
    Minimum 12 mm per column.
    """
    n = len(headers)
    if n == 0:
        return []
    max_lens = [len(str(h)) for h in headers]
    for row in rows[:max_sample]:
        for i, cell in enumerate(row[:n]):
            max_lens[i] = max(max_lens[i], len(str(cell)))
    total  = sum(max_lens) or 1
    widths = [max(12.0, page_width * (ml / total)) for ml in max_lens]
    scale  = page_width / sum(widths)
    return [w * scale for w in widths]


# -- Word-wrap helper ----------------------------------------------------------

def _wrap_text(pdf, text: str, col_width_mm: float) -> list[str]:
    """
    Split text into lines that fit within col_width_mm using pdf.get_string_width().
    Handles both natural word boundaries and hard wrapping for words wider than
    the column.
    """
    words  = text.split(" ")
    lines  = []
    current = ""

    for word in words:
        candidate = (current + " " + word).strip() if current else word
        if pdf.get_string_width(candidate) <= col_width_mm:
            current = candidate
        else:
            if current:
                lines.append(current)
            # Hard-wrap the word itself if it's wider than the column
            while pdf.get_string_width(word) > col_width_mm:
                for cut in range(len(word), 0, -1):
                    if pdf.get_string_width(word[:cut]) <= col_width_mm:
                        lines.append(word[:cut])
                        word = word[cut:]
                        break
                else:
                    lines.append(word)
                    word = ""
                    break
            current = word

    if current:
        lines.append(current)
    return lines or [""]


# -- PDF builder ---------------------------------------------------------------

def convert_to_pdf(
    input_path: Path,
    output_path: Path,
    title: str = "Wazuh Report",
    subtitle: str = "",
    report_date: str = "",
    logo_path: Path | None = None,
) -> Path:
    """
    Convert a CSV or XLSX file to a formatted PDF.

    Each row is measured before drawing so its height exactly fits its
    content with no inter-row gaps.

    Parameters
    ----------
    input_path  : source file (.xlsx or .csv)
    output_path : destination .pdf file
    title       : report title on every page
    subtitle    : optional second line under the title
    report_date : date string in header (defaults to now)
    logo_path   : path to a PNG logo placed top-right of the header.
                  Defaults to assets/wazuh-logo.png next to this script.
                  Pass logo_path=False to suppress the logo entirely.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("pip install fpdf2")

    logging.info(f"  Converting {input_path.name} -> {output_path.name}")
    headers, rows = load_file(input_path)
    if not headers:
        raise ValueError(f"'{input_path.name}' is empty.")
    if not report_date:
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Resolve logo path -- fall back to assets/wazuh-logo.png if not given
    if logo_path is None:
        _default = Path(__file__).resolve().parent.parent / "assets" / "wazuh-logo.png"
        logo_path = _default if _default.exists() else None
    elif logo_path is False:
        logo_path = None
    # Logo print dimensions (proportional to 245x78 source at target 35mm wide)
    LOGO_W = 35.0   # mm
    LOGO_H = 11.1   # mm

    FONT_SIZE   = 8.0
    TITLE_SIZE  = 14.0
    SUBTITLE_SZ = 9.0
    HEADER_H    = 7.0    # fixed height for column-label row
    LINE_H      = 5.0    # height of a single text line in data cells
    CELL_PAD    = 1.5    # left padding (mm) subtracted from usable cell width

    class WazuhPDF(FPDF):
        def footer(self):
            self.set_y(-10)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*FOOTER_FG)
            self.cell(
                0, 5,
                _safe(f"Wazuh Report System  |  Page {self.page_no()} of {{nb}}"
                      f"  |  {report_date}"),
                align="C",
            )

    pdf = WazuhPDF(orientation="L", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=False)   # we handle page breaks manually
    pdf.set_margins(left=10, top=10, right=10)
    pdf.add_page()

    USABLE_W  = pdf.w - pdf.l_margin - pdf.r_margin
    PAGE_H    = pdf.h
    FOOT_H    = 15        # space reserved for footer
    col_widths = calculate_col_widths(headers, rows, USABLE_W, FONT_SIZE)

    # -- Page header -----------------------------------------------------------
    def draw_page_header() -> float:
        """Draw title block + column labels. Returns Y position after header."""
        pdf.set_xy(pdf.l_margin, pdf.t_margin)

        # Logo: top-right corner of the header
        if logo_path and Path(logo_path).exists():
            pdf.image(
                str(logo_path),
                x=pdf.w - pdf.r_margin - LOGO_W,
                y=pdf.t_margin,
                w=LOGO_W,
                h=LOGO_H,
            )

        # Title: constrain width so it never overlaps the logo
        _title_w = (USABLE_W - LOGO_W - 5) if (logo_path and Path(logo_path).exists()) else USABLE_W
        pdf.set_font("Helvetica", "B", TITLE_SIZE)
        pdf.set_text_color(*TITLE_FG)
        pdf.cell(_title_w, 8, _safe(title), align="L", new_x="LMARGIN", new_y="NEXT")

        if subtitle:
            pdf.set_font("Helvetica", "", SUBTITLE_SZ)
            pdf.set_text_color(*SUBTITLE_FG)
            pdf.cell(0, 5, _safe(subtitle), align="L", new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_text_color(*SUBTITLE_FG)
        pdf.cell(0, 5, _safe(f"Generated: {report_date}"), align="L", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Column labels
        pdf.set_font("Helvetica", "B", FONT_SIZE)
        pdf.set_fill_color(*HEADER_BG)
        pdf.set_text_color(*HEADER_FG)
        pdf.set_draw_color(*BORDER)
        pdf.set_line_width(0.2)
        for h, w in zip(headers, col_widths):
            pdf.cell(w, HEADER_H, _safe(str(h)), border=1, fill=True, align="L")
        pdf.ln()
        return pdf.get_y()

    y = draw_page_header()

    # -- Data rows -------------------------------------------------------------
    pdf.set_font("Helvetica", "", FONT_SIZE)

    for row_idx, row in enumerate(rows):
        cells = [_safe(str(row[i]) if i < len(row) else "") for i in range(len(col_widths))]

        # Pass 1 -- measure: find the maximum number of lines across all cells
        wrapped = []
        for val, w in zip(cells, col_widths):
            lines = _wrap_text(pdf, val, w - CELL_PAD)
            wrapped.append(lines)

        max_lines  = max(len(ln) for ln in wrapped)
        row_height = max_lines * LINE_H

        # Page break: does this row fit before the footer?
        if y + row_height > PAGE_H - FOOT_H:
            pdf.add_page()
            y = draw_page_header()
            pdf.set_font("Helvetica", "", FONT_SIZE)
            pdf.set_draw_color(*BORDER)
            pdf.set_line_width(0.2)

        fill_color = ROW_EVEN if row_idx % 2 == 0 else ROW_ODD

        # Pass 2 -- draw: each cell rendered as stacked line sub-cells
        x = pdf.l_margin
        for lines, w in zip(wrapped, col_widths):
            cy = y
            pdf.set_fill_color(*fill_color)
            pdf.set_text_color(30, 30, 30)
            pdf.set_draw_color(*BORDER)
            pdf.set_line_width(0.2)

            for line_idx, line in enumerate(lines):
                # Top border only on first line, bottom border only on last
                if len(lines) == 1:
                    border = 1          # all sides
                elif line_idx == 0:
                    border = "LTR"      # top of cell
                elif line_idx == len(lines) - 1:
                    border = "LBR"      # bottom of cell
                else:
                    border = "LR"       # sides only

                pdf.set_xy(x, cy)
                pdf.cell(w, LINE_H, line, border=border, fill=True, align="L")
                cy += LINE_H

            # If this cell has fewer lines than the tallest, pad with blank cells
            while cy < y + row_height:
                remaining = y + row_height - cy
                pad_h     = min(LINE_H, remaining)
                is_last   = (cy + pad_h >= y + row_height)
                border    = "LBR" if is_last else "LR"
                pdf.set_xy(x, cy)
                pdf.cell(w, pad_h, "", border=border, fill=True)
                cy += pad_h

            x += w

        y += row_height

    # -- Write -----------------------------------------------------------------
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(output_path))
    logging.info(f"  PDF written: {output_path} ({output_path.stat().st_size:,} bytes)")
    return output_path


# -- CLI -----------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Convert XLSX/CSV to PDF")
    parser.add_argument("input",      help="Input file (.xlsx or .csv)")
    parser.add_argument("output",     help="Output PDF file")
    parser.add_argument("--title",    default="Wazuh Report")
    parser.add_argument("--subtitle", default="")
    parser.add_argument("--date",     default="")
    parser.add_argument("--logo",     default=None,
                        help="Path to logo PNG (default: assets/wazuh-logo.png)")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    convert_to_pdf(
        input_path=Path(args.input),
        output_path=Path(args.output),
        title=args.title,
        subtitle=args.subtitle,
        report_date=args.date,
        logo_path=Path(args.logo) if args.logo else None,
    )
    print(f"Done: {args.output}")


if __name__ == "__main__":
    main()
